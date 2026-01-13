"""
Task report generation service
Generates execution reports in Excel, PDF, and JSON formats
"""
import logging
import os
import json
from typing import Dict, Any, Optional
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app import db
from app.models.link import ImportTask, Link
from app.services.task_service import get_task_service

logger = logging.getLogger(__name__)


class TaskReportService:
    """Service for generating task execution reports"""

    def __init__(self):
        self.reports_dir = Path(os.getenv('REPORTS_DIR', 'reports'))
        self.reports_dir.mkdir(exist_ok=True, parents=True)

    def generate_report(self, task_id: int,
                       format_type: str = 'excel') -> Dict[str, Any]:
        """
        Generate task execution report

        Args:
            task_id: Task ID
            format_type: 'excel', 'pdf', or 'json'

        Returns:
            Dict with success status and filepath
        """
        try:
            task_service = get_task_service()
            task_summary = task_service.get_task_summary(task_id)

            if not task_summary:
                return {
                    'success': False,
                    'error': f'Task {task_id} not found'
                }

            if format_type == 'excel':
                filepath = self._generate_excel_report(task_id, task_summary)
            elif format_type == 'pdf':
                filepath = self._generate_pdf_report(task_id, task_summary)
            elif format_type == 'json':
                filepath = self._generate_json_report(task_id, task_summary)
            else:
                return {
                    'success': False,
                    'error': f'Unsupported format: {format_type}'
                }

            logger.info(f"Generated {format_type} report for task {task_id}: {filepath}")

            return {
                'success': True,
                'filepath': str(filepath),
                'format': format_type
            }

        except Exception as e:
            logger.error(f"Failed to generate report for task {task_id}: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }

    def _generate_excel_report(self, task_id: int, task_summary: Dict) -> Path:
        """Generate Excel report with formatted sheets"""
        task = task_summary

        # Get task links for details
        task_service = get_task_service()
        links = task_service.get_task_links(task_id)

        # Create workbook
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header styling
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        label_font = Font(bold=True, size=10)

        # Summary Sheet Title
        ws_summary['A1'] = 'Task Execution Report'
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:B1')

        # Summary data
        row = 3
        summary_data = [
            ('Task ID', task['id']),
            ('Task Name', task['name']),
            ('Status', task['status']),
            ('Created At', task.get('created_at').isoformat() if task.get('created_at') else 'N/A'),
            ('Started At', task.get('started_at').isoformat() if task.get('started_at') else 'N/A'),
            ('Completed At', task.get('completed_at').isoformat() if task.get('completed_at') else 'N/A'),
            ('', ''),
            ('Total Links', task.get('total_links', 0)),
            ('Processed Links', task.get('processed_links', 0)),
        ]

        # Add link statistics if available
        if 'link_statistics' in task:
            stats = task['link_statistics']
            summary_data.extend([
                ('', ''),
                ('Link Statistics', ''),
                ('  Valid Links', stats.get('valid', 0)),
                ('  Invalid Links', stats.get('invalid', 0)),
                ('  Pending Validation', stats.get('pending_validation', 0)),
            ])

        for label, value in summary_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            if label and not label.startswith('  '):
                ws_summary[f'A{row}'].font = label_font
            row += 1

        # Auto-adjust column widths for summary
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40

        # Links Details Sheet
        if links:
            ws_links = wb.create_sheet("Links Details")

            # Headers
            headers = ['ID', 'Title', 'URL', 'Source', 'Valid', 'Validation Status', 'Imported At']
            for col, header in enumerate(headers, 1):
                cell = ws_links.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data rows
            for row_idx, link in enumerate(links, 2):
                ws_links.cell(row=row_idx, column=1, value=link.id)
                ws_links.cell(row=row_idx, column=2, value=link.title or 'N/A')
                ws_links.cell(row=row_idx, column=3, value=link.url)
                ws_links.cell(row=row_idx, column=4, value=link.source or 'N/A')
                ws_links.cell(row=row_idx, column=5, value='Yes' if link.is_valid else 'No' if link.is_valid is False else 'Pending')
                ws_links.cell(row=row_idx, column=6, value=link.validation_status or 'N/A')
                ws_links.cell(row=row_idx, column=7, value=link.imported_at.isoformat() if link.imported_at else 'N/A')

            # Auto-adjust column widths for links
            ws_links.column_dimensions['A'].width = 8
            ws_links.column_dimensions['B'].width = 40
            ws_links.column_dimensions['C'].width = 50
            ws_links.column_dimensions['D'].width = 12
            ws_links.column_dimensions['E'].width = 10
            ws_links.column_dimensions['F'].width = 18
            ws_links.column_dimensions['G'].width = 25

        # Save file
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"task_{task['id']}_report_{timestamp}.xlsx"
        filepath = self.reports_dir / filename

        wb.save(filepath)
        logger.info(f"Generated Excel report: {filepath}")
        return filepath

    def _generate_pdf_report(self, task_id: int, task_summary: Dict) -> Path:
        """
        Generate PDF report (placeholder - requires reportlab)

        Currently falls back to JSON format.
        To implement full PDF support, install reportlab and implement PDF generation.
        """
        logger.warning("PDF generation not yet implemented, falling back to JSON")
        return self._generate_json_report(task_id, task_summary)

    def _generate_json_report(self, task_id: int, task_summary: Dict) -> Path:
        """Generate JSON report"""
        # Get task links for full report
        task_service = get_task_service()
        links = task_service.get_task_links(task_id)

        # Build comprehensive report
        report_data = {
            'report_generated_at': datetime.utcnow().isoformat(),
            'task': {
                'id': task_summary['id'],
                'name': task_summary['name'],
                'status': task_summary['status'],
                'total_links': task_summary.get('total_links', 0),
                'processed_links': task_summary.get('processed_links', 0),
                'config': task_summary.get('config', {}),
                'created_at': task_summary.get('created_at').isoformat() if task_summary.get('created_at') else None,
                'started_at': task_summary.get('started_at').isoformat() if task_summary.get('started_at') else None,
                'completed_at': task_summary.get('completed_at').isoformat() if task_summary.get('completed_at') else None,
                'statistics': task_summary.get('link_statistics', {})
            },
            'links': [
                {
                    'id': link.id,
                    'title': link.title,
                    'url': link.url,
                    'source': link.source,
                    'is_valid': link.is_valid,
                    'validation_status': link.validation_status,
                    'imported_at': link.imported_at.isoformat() if link.imported_at else None,
                    'priority': link.priority
                }
                for link in links
            ]
        }

        # Save to file
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"task_{task_summary['id']}_report_{timestamp}.json"
        filepath = self.reports_dir / filename

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

        logger.info(f"Generated JSON report: {filepath}")
        return filepath

    def cleanup_old_reports(self, days: int = 30) -> int:
        """
        Clean up reports older than specified days

        Args:
            days: Number of days to keep reports

        Returns:
            Number of deleted reports
        """
        try:
            cutoff_date = datetime.utcnow().timestamp() - (days * 24 * 60 * 60)
            deleted_count = 0

            for report_file in self.reports_dir.glob('task_*_report_*'):
                if report_file.stat().st_mtime < cutoff_date:
                    report_file.unlink()
                    deleted_count += 1
                    logger.info(f"Deleted old report: {report_file.name}")

            logger.info(f"Cleaned up {deleted_count} old reports")
            return deleted_count

        except Exception as e:
            logger.error(f"Failed to cleanup old reports: {e}", exc_info=True)
            return 0


# Singleton instance
_task_report_service = None


def get_task_report_service() -> TaskReportService:
    """Get singleton instance of TaskReportService"""
    global _task_report_service
    if _task_report_service is None:
        _task_report_service = TaskReportService()
    return _task_report_service
